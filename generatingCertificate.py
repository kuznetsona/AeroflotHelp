import sys
from PyQt5 import QtWidgets
from openpyxl import load_workbook
from ui_generatingCertificate_v2 import Ui_MainWindow
from PyQt5.QtGui import QRegExpValidator
from PyQt5.QtCore import QRegExp
from docx import Document
import shutil

class CertificateApp(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.TypeCertificate.clear()
        self.TypeCertificate.addItems(["Справка1", "Справка2", "Справка3", "Справка4"])

        self.telephoneLineEdit.setInputMask("+7 (999) 999-9999;_")
        self.passportLineEdit.setInputMask("99 99 999999;_")
        self.ticketLineEdit.setInputMask("999999")

        emailRegex = QRegExp("[a-z0-9._%+-]+@[a-z0-9.-]+\.[a-z]{2,}")
        emailValidator = QRegExpValidator(emailRegex)
        self.emailLineEdit.setValidator(emailValidator)

        nameRegex = QRegExp("[A-Za-zа-яА-ЯёЁ-]+")
        nameValidator = QRegExpValidator(nameRegex)
        self.nameLineEdit.setValidator(nameValidator)
        self.lastNameLineEdit.setValidator(nameValidator)
        self.surnameLineEdit.setValidator(nameValidator)
        # self.nameRecipient.setValidator(nameValidator)

        self.nameLineEdit.textEdited.connect(self.capitalizeFirstLetter)
        self.lastNameLineEdit.textEdited.connect(self.capitalizeFirstLetter)
        self.surnameLineEdit.textEdited.connect(self.capitalizeFirstLetter)
        # self.nameRecipient.textEdited.connect(self.capitalizeFirstLetter)

        self.PreviewButton.clicked.connect(self.printCertificate)



    # Инициализация и запуск приложения

    def capitalizeFirstLetter(self, text):
        if text:
            # Преобразование первого символа в верхний регистр
            capitalized_text = text[0].upper() + text[1:]
            # Изменение текста в виджете QLineEdit без повторного вызова события textEdited
            self.sender().blockSignals(True)
            self.sender().setText(capitalized_text)
            self.sender().blockSignals(False)


    def printCertificate(self):
        # Загружаем книгу Excel и активный лист
        workbook = load_workbook('database.xlsx')
        sheet = workbook.active

        # Словарь заголовков и соответствующих значений для записи
        data = {
            'TypeCertificate': self.TypeCertificate.currentText(),
            'NumberTicket': self.ticketLineEdit.text(),
            'Itinerary': self.Itinerary.text(),
            'LastName': self.lastNameLineEdit.text(),
            'Name': self.nameLineEdit.text(),
            'Surname': self.surnameLineEdit.text(),
            'Passport': self.passportLineEdit.text(),
            'IssuingAuthority': self.PassportInformation.text(),
            'Telephone': self.telephoneLineEdit.text(),
            'Email': self.emailLineEdit.text(),
            'RecipientName': self.nameRecipient.text(),
        }

        # Сопоставляем колонки Excel и данные с формы
        headers = [cell.value for cell in sheet[1]]  # Предполагаем, что заголовки находятся в первой строке
        next_row = sheet.max_row + 1
        for header in headers:
            col = headers.index(header) + 1  # Получаем индекс заголовка для определения номера столбца
            sheet.cell(row=next_row, column=col, value=data[header])

        # Сохраняем изменения
        workbook.save('database.xlsx')
        QtWidgets.QMessageBox.information(self, 'Успех', 'Данные добавлены в файл database.xlsx')



        # Создаём копию шаблона Word
        template_path = 'factOfFlight.docx'
        new_doc_path = 'updated_document.docx'
        shutil.copy(template_path, new_doc_path)

        # Загружаем документ Word (копию)
        doc = Document(new_doc_path)

        # Извлекаем данные из последней добавленной строки в Excel
        last_row = sheet.max_row
        data_row = {sheet.cell(row=1, column=col).value: sheet.cell(row=last_row, column=col).value for col in
                    range(1, sheet.max_column + 1)}

        # Вставляем данные в документ Word
        for paragraph in doc.paragraphs:
            for key, value in data_row.items():
                if value and f'{{{{{key}}}}}' in paragraph.text:
                    # Обновляем текст абзаца, сохраняя форматирование
                    runs = list(paragraph.runs)
                    for i in range(len(runs)):
                        if f'{{{{{key}}}}}' in runs[i].text:
                            runs[i].text = runs[i].text.replace(f'{{{{{key}}}}}', str(value))

        # Сохраняем документ
        doc.save(new_doc_path)

        # Сообщение об успешном выполнении
        QtWidgets.QMessageBox.information(self, 'Успех',
                                          'Данные добавлены в файлы database.xlsx и updated_document.docx')




# Инициализация и запуск приложения
app = QtWidgets.QApplication(sys.argv)
mainWindow = CertificateApp()
mainWindow.show()
sys.exit(app.exec_())
